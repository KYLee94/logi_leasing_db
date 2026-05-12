#!/usr/bin/env python
"""Build logistics source extracts and optional xlsx-vs-Sheets diffs.

The output is intentionally cell-level. It is the migration evidence layer for
public.ll_source_* and must stay independent from existing Supabase tables.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT_DIR = ROOT / "qa-artifacts" / "source-diff"

SHEET_CONFIG: Dict[str, Dict[str, int]] = {
    "Meta_데이터 항목 설명": {"header_row": 2, "data_start_row": 3},
    "DB_일반": {"header_row": 9, "data_start_row": 12},
    "DB_히스토리 누적": {"header_row": 10, "data_start_row": 15},
    "Log": {"header_row": 4, "data_start_row": 5},
    "자산_담당자 연결": {"header_row": 3, "data_start_row": 4},
}


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def make_hash(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=json_default)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def normalized_header(value: Any, column_index: int) -> str:
    text = normalize_cell_value(value).strip().lower()
    keep = []
    for ch in text:
        if ch.isalnum() or "\uac00" <= ch <= "\ud7a3":
            keep.append(ch)
        else:
            keep.append("_")
    result = "".join(keep).strip("_")
    while "__" in result:
        result = result.replace("__", "_")
    return result[:80] or f"blank_col_{column_index:03d}"


def infer_value_type(value: Any, formula: str) -> str:
    if formula:
        return "formula"
    if value is None or normalize_cell_value(value).strip() == "":
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    if isinstance(value, (datetime, date)):
        return "date"
    return "text"


def find_xlsx(path_arg: Optional[str]) -> Path:
    if path_arg:
        path = Path(path_arg)
        if not path.is_absolute():
            path = ROOT / path
        if not path.exists():
            raise FileNotFoundError(path)
        return path
    candidates = [p for p in ROOT.glob("*.xlsx") if "260414" in p.name]
    if not candidates:
        raise FileNotFoundError("No xlsx file containing 260414 found in workspace root.")
    return candidates[0]


def used_range(ws: Any) -> Tuple[int, int]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    return max_row, max_col


def extract_workbook(xlsx_path: Path) -> Dict[str, Any]:
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    extracted_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    workbook = {
        "source_kind": "xlsx",
        "source_system": "google_sheets",
        "source_name": xlsx_path.name,
        "source_path": str(xlsx_path),
        "extracted_at": extracted_at,
        "sheets": [],
        "columns": [],
        "rows": [],
        "cells": [],
        "summary": {
            "sheet_count": 0,
            "row_count": 0,
            "column_count": 0,
            "cell_count": 0,
            "formula_cell_count": 0,
            "non_empty_cell_count": 0,
        },
    }

    for sheet_index, sheet_name in enumerate(wb_formula.sheetnames, start=1):
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        max_row, max_col = used_range(ws_formula)
        cfg = SHEET_CONFIG.get(sheet_name, {})
        header_row = cfg.get("header_row", 1)
        sheet_id = f"xlsx_{sheet_index:02d}_{make_hash(sheet_name)[:10]}"
        sheet_cell_count = max_row * max_col
        workbook["sheets"].append(
            {
                "sheet_id": sheet_id,
                "import_id": None,
                "sheet_name": sheet_name,
                "source_file": xlsx_path.name,
                "row_count": max_row,
                "column_count": max_col,
                "cell_count": sheet_cell_count,
                "header_row": header_row,
                "data_start_row": cfg.get("data_start_row", header_row + 1),
                "source_hash": make_hash({"sheet": sheet_name, "rows": max_row, "cols": max_col}),
            }
        )

        column_ids = {}
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header_value = ws_values.cell(header_row, col_idx).value if header_row <= max_row else None
            column_id = f"{sheet_id}_c{col_idx:04d}"
            column_ids[col_idx] = column_id
            workbook["columns"].append(
                {
                    "column_id": column_id,
                    "sheet_id": sheet_id,
                    "column_index": col_idx,
                    "column_letter": col_letter,
                    "header_value": normalize_cell_value(header_value),
                    "normalized_header": normalized_header(header_value, col_idx),
                    "column_role": "business_value" if normalize_cell_value(header_value).strip() else "blank_header",
                }
            )

        for row_idx in range(1, max_row + 1):
            row_values: List[str] = []
            non_empty = 0
            row_id = f"{sheet_id}_r{row_idx:06d}"
            for col_idx in range(1, max_col + 1):
                formula_cell = ws_formula.cell(row_idx, col_idx)
                value_cell = ws_values.cell(row_idx, col_idx)
                formula_text = ""
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                    formula_text = formula_cell.value
                raw_value = value_cell.value if formula_text else formula_cell.value
                display_value = normalize_cell_value(raw_value)
                if display_value.strip() or formula_text:
                    non_empty += 1
                row_values.append(display_value + ("|" + formula_text if formula_text else ""))

                col_letter = get_column_letter(col_idx)
                a1_ref = f"{sheet_name}!{col_letter}{row_idx}"
                note = formula_cell.comment.text if formula_cell.comment else ""
                cell_payload = {
                    "sheet_name": sheet_name,
                    "row_number": row_idx,
                    "column_index": col_idx,
                    "column_letter": col_letter,
                    "a1_ref": a1_ref,
                    "display_value": display_value,
                    "raw_value": normalize_cell_value(raw_value),
                    "formula": formula_text,
                    "is_blank": not bool(display_value.strip() or formula_text),
                    "number_format": normalize_cell_value(formula_cell.number_format),
                    "note": note,
                    "value_type": infer_value_type(raw_value, formula_text),
                }
                cell_payload["source_hash"] = make_hash(cell_payload)
                cell_payload.update(
                    {
                        "cell_id": f"{row_id}_c{col_idx:04d}",
                        "row_id": row_id,
                        "column_id": column_ids[col_idx],
                    }
                )
                workbook["cells"].append(cell_payload)

            row_hash = make_hash(row_values)
            workbook["rows"].append(
                {
                    "row_id": row_id,
                    "sheet_id": sheet_id,
                    "row_number": row_idx,
                    "row_hash": row_hash,
                    "non_empty_cell_count": non_empty,
                }
            )

        workbook["summary"]["row_count"] += max_row
        workbook["summary"]["column_count"] += max_col
        workbook["summary"]["cell_count"] += sheet_cell_count

    workbook["summary"]["sheet_count"] = len(workbook["sheets"])
    workbook["summary"]["formula_cell_count"] = sum(1 for cell in workbook["cells"] if cell["formula"])
    workbook["summary"]["non_empty_cell_count"] = sum(1 for cell in workbook["cells"] if not cell["is_blank"])
    return workbook


def index_cells(extract: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    return {cell["a1_ref"]: cell for cell in extract.get("cells", [])}


def build_diff(xlsx_extract: Dict[str, Any], sheet_extract: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not sheet_extract:
        return {
            "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
            "status": "sheet_extract_missing",
            "summary": {
                "match": 0,
                "value_diff": 0,
                "formula_diff": 0,
                "xlsx_only": 0,
                "sheet_only": 0,
                "blank_diff": 0,
            },
            "diffs": [],
        }
    xlsx_cells = index_cells(xlsx_extract)
    sheet_cells = index_cells(sheet_extract)
    all_refs = sorted(set(xlsx_cells.keys()) | set(sheet_cells.keys()))
    summary = {key: 0 for key in ["match", "value_diff", "formula_diff", "xlsx_only", "sheet_only", "blank_diff"]}
    diffs = []
    for ref in all_refs:
        x_cell = xlsx_cells.get(ref)
        s_cell = sheet_cells.get(ref)
        if not x_cell:
            diff_type = "sheet_only"
        elif not s_cell:
            diff_type = "xlsx_only"
        elif bool(x_cell.get("is_blank")) != bool(s_cell.get("is_blank")):
            diff_type = "blank_diff"
        elif normalize_cell_value(x_cell.get("formula")) != normalize_cell_value(s_cell.get("formula")):
            diff_type = "formula_diff"
        elif normalize_cell_value(x_cell.get("display_value")) != normalize_cell_value(s_cell.get("display_value")):
            diff_type = "value_diff"
        else:
            diff_type = "match"
        summary[diff_type] += 1
        if diff_type != "match":
            diffs.append(
                {
                    "diff_id": make_hash({"a1_ref": ref, "diff_type": diff_type}),
                    "a1_ref": ref,
                    "diff_type": diff_type,
                    "xlsx_value": normalize_cell_value(x_cell.get("display_value") if x_cell else ""),
                    "sheet_value": normalize_cell_value(s_cell.get("display_value") if s_cell else ""),
                    "xlsx_formula": normalize_cell_value(x_cell.get("formula") if x_cell else ""),
                    "sheet_formula": normalize_cell_value(s_cell.get("formula") if s_cell else ""),
                    "chosen_source": "google_sheets" if s_cell else "xlsx",
                }
            )
    return {
        "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "status": "ok",
        "summary": summary,
        "diffs": diffs,
    }


def write_csv(path: Path, rows: Iterable[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=None)
    parser.add_argument("--sheet-extract", default=None)
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_extract = extract_workbook(find_xlsx(args.xlsx))
    sheet_extract = None
    if args.sheet_extract:
        with Path(args.sheet_extract).open("r", encoding="utf-8") as handle:
            sheet_extract = json.load(handle)
    diff = build_diff(xlsx_extract, sheet_extract)

    summary = {
        "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "xlsx": xlsx_extract["summary"],
        "sheet": sheet_extract.get("summary") if sheet_extract else None,
        "diff": diff["summary"],
        "diff_status": diff["status"],
        "adoption_rule": "live_google_sheets_first; xlsx_only metadata/source is preserved in ll_source_*",
        "required_source_fields": [
            "sheet_name",
            "row_number",
            "column_index",
            "column_letter",
            "a1_ref",
            "display_value",
            "raw_value",
            "formula",
            "is_blank",
            "number_format",
            "note",
            "source_hash",
        ],
    }

    (out_dir / "xlsx-extract.json").write_text(json.dumps(xlsx_extract, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    (out_dir / "source-diff.json").write_text(json.dumps(diff, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    (out_dir / "source-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    write_csv(
        out_dir / "source-diff.csv",
        diff["diffs"],
        ["diff_id", "a1_ref", "diff_type", "xlsx_value", "sheet_value", "xlsx_formula", "sheet_formula", "chosen_source"],
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
