#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Apply supplemental logistics asset data from the 2026-05 Excel workbooks.

This script is intentionally deterministic. It updates the public GitHub Pages
fallback payloads and prepares a Supabase ll_* supplemental dataset, without
calling external services.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps


ROOT = Path(__file__).resolve().parents[2]
DOCS_DATA = ROOT / "docs" / "data"
DOCS_STAFF = ROOT / "docs" / "staff"
ARTIFACT_DIR = ROOT / "qa-artifacts" / "supabase"

PERMISSION_WORKBOOK_PREFIX = "260513_"
PERMISSION_WORKBOOK_SUFFIX = "_수식 제거.xlsx"
FUND_WORKBOOK_PREFIX = "260520_"

GENERATED_AT = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
PY_PER_SQM = 1 / 3.305785

NEW_ASSET_CODES = ["A190002001", "A190013001"]

ASSET_OVERRIDES: dict[str, dict[str, Any]] = {
    "A190002001": {
        "assetName": "분당야탑물류센터",
        "address": "경기도 성남시 분당구 탄천로 257",
        "parcelAddress": "경기도 성남시 분당구 야탑동 403",
        "latitude": None,
        "longitude": None,
        "coordinateStatus": "좌표 재확인 필요",
        "coordinateNote": "기존 좌표 37.40971,127.12046은 OpenStreetMap 역지오코딩 기준 탄천종합운동장으로 확인되어 핀 표시에서 제외합니다.",
        "buildingRegisterStatus": "pending_edge_function_lookup",
        "buildingRegisterNote": "성남시 분당구 야탑동 403 기준으로 Edge Function에서 조회합니다.",
        "buildingRegisterQuery": {
            "sigunguCd": "41135",
            "bjdongCd": "10700",
            "platGbCd": "0",
            "bun": "0403",
            "ji": "0000",
        },
    },
    "A190013001": {
        "assetName": "포천정교리물류센터",
        "address": "경기도 포천시 가산면 정교리 272-1",
        "parcelAddress": "경기도 포천시 가산면 정교리 272-1",
        "latitude": 37.83925,
        "longitude": 127.18528,
        "coordinateStatus": "주소 기반 좌표",
        "coordinateNote": "개발 진행 중 자산으로 건축물대장/지번 기반 추가 검증이 필요합니다.",
        "buildingRegisterStatus": "development_asset_not_found_expected",
        "buildingRegisterNote": "개발 진행 중 자산으로 건축물대장 미존재 가능성을 정상 상태로 기록합니다.",
        "buildingRegisterQuery": {
            "sigunguCd": "41650",
            "bjdongCd": "",
            "platGbCd": "0",
            "bun": "0272",
            "ji": "0001",
        },
    },
}


def read_json(path: Path, fallback: Any = None) -> Any:
    if not path.exists():
        return fallback
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\xa0", " ").strip()


def yes(value: Any) -> bool:
    return clean(value).upper() == "Y"


def date_text(value: Any) -> str:
    return clean(value)


def stable_hash(value: Any) -> str:
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def safe_id(value: str) -> str:
    return re.sub(r"[^a-z0-9가-힣_]+", "_", value.lower()).strip("_") or "unknown"


def asset_id(asset_code: str) -> str:
    return "asset_" + asset_code.lower()


def fund_id(fund_code: str) -> str:
    return "fund_" + safe_id(fund_code)


def permission_id(email: str, scope_type: str, scope_id: str) -> str:
    return "perm_" + safe_id(email.replace("@", "_at_")) + "_" + safe_id(scope_type) + "_" + safe_id(scope_id)


def amount(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def find_workbook(prefix: str, suffix: str | None = None) -> Path:
    candidates = sorted(ROOT.glob(f"{prefix}*.xlsx"))
    if suffix:
        candidates = [path for path in candidates if path.name.endswith(suffix)]
    if not candidates:
        raise FileNotFoundError(f"Workbook not found: {prefix} {suffix or ''}")
    return candidates[0]


def parse_permission_workbook() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    path = find_workbook(PERMISSION_WORKBOOK_PREFIX, PERMISSION_WORKBOOK_SUFFIX)
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    people: list[dict[str, Any]] = []
    managers: list[dict[str, Any]] = []
    asset_header_row = None

    for row in range(4, ws.max_row + 1):
        col_b = clean(ws.cell(row, 2).value)
        if col_b == "자산코드":
            asset_header_row = row
            break
        if not col_b or col_b.startswith("국내 "):
            continue
        asset_codes = [item.strip() for item in clean(ws.cell(row, 13).value).split(",") if item.strip()]
        people.append({
            "sourceWorkbook": path.name,
            "sourceRow": row,
            "staffName": col_b,
            "email": clean(ws.cell(row, 3).value).replace("\t", ""),
            "organization": clean(ws.cell(row, 4).value),
            "assignedCanRead": yes(ws.cell(row, 5).value),
            "assignedCanCreate": yes(ws.cell(row, 6).value),
            "assignedCanUpdate": yes(ws.cell(row, 7).value),
            "assignedCanDelete": yes(ws.cell(row, 8).value),
            "otherCanRead": yes(ws.cell(row, 9).value),
            "otherCanCreate": yes(ws.cell(row, 10).value),
            "otherCanUpdate": yes(ws.cell(row, 11).value),
            "otherCanDelete": yes(ws.cell(row, 12).value),
            "assetCodes": asset_codes,
        })

    if asset_header_row is None:
        raise ValueError("Permission workbook asset table header was not found.")

    for row in range(asset_header_row + 1, ws.max_row + 1):
        code = clean(ws.cell(row, 2).value)
        if not code:
            continue
        managers.append({
            "sourceWorkbook": path.name,
            "sourceRow": row,
            "assetCode": code,
            "assetId": asset_id(code),
            "assetName": clean(ws.cell(row, 3).value),
            "fundCode": clean(ws.cell(row, 4).value),
            "fundName": clean(ws.cell(row, 5).value),
            "managerName": clean(ws.cell(row, 6).value),
            "organization": clean(ws.cell(row, 7).value),
            "email": clean(ws.cell(row, 8).value),
        })
    return people, managers


def rows_from_sheet(ws, start_row: int = 4) -> list[dict[str, Any]]:
    headers = [clean(ws.cell(3, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row in range(start_row, ws.max_row + 1):
        values = {headers[col - 1]: ws.cell(row, col).value for col in range(1, ws.max_column + 1) if headers[col - 1]}
        if not any(value is not None and clean(value) for value in values.values()):
            continue
        values["_sourceRow"] = row
        rows.append(values)
    return rows


def parse_fund_workbook() -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    path = find_workbook(FUND_WORKBOOK_PREFIX)
    wb = load_workbook(path, data_only=True)
    fund_rows = rows_from_sheet(wb.worksheets[0])
    beneficiary_rows = rows_from_sheet(wb.worksheets[1])
    lender_rows = rows_from_sheet(wb.worksheets[2])

    funds: dict[str, dict[str, Any]] = {}
    for row in fund_rows:
        code = clean(row.get("자산코드"))
        if not code:
            continue
        funds[code] = {
            "sourceWorkbook": path.name,
            "sourceRow": row["_sourceRow"],
            "assetCode": code,
            "assetId": asset_id(code),
            "assetName": clean(row.get("자산명")),
            "fundCode": clean(row.get("펀드코드")),
            "fundName": clean(row.get("펀드명")),
            "shortName": clean(row.get("약칭")),
            "legalForm": clean(row.get("법적형태")),
            "sector": clean(row.get("투자섹터")),
            "fundType": clean(row.get("펀드유형")),
            "investmentStrategy": clean(row.get("투자전략")),
            "setupDate": date_text(row.get("최초설정일")),
            "maturityDate": date_text(row.get("만기일")),
        }

    beneficiaries: list[dict[str, Any]] = []
    for row in beneficiary_rows:
        code = clean(row.get("자산코드"))
        name = clean(row.get("수익자"))
        if not code or not name:
            continue
        beneficiaries.append({
            "sourceWorkbook": path.name,
            "sourceRow": row["_sourceRow"],
            "assetCode": code,
            "assetId": asset_id(code),
            "assetName": clean(row.get("자산명")),
            "fundCode": clean(row.get("펀드코드")),
            "fundName": clean(row.get("펀드명")),
            "tranche": clean(row.get("Tranche")),
            "beneficiaryName": name,
            "investmentAmountKrw": amount(row.get("투입금액(원)")),
        })

    lenders: list[dict[str, Any]] = []
    for row in lender_rows:
        code = clean(row.get("자산코드"))
        lender = clean(row.get("대주"))
        if not code or not lender:
            continue
        lenders.append({
            "sourceWorkbook": path.name,
            "sourceRow": row["_sourceRow"],
            "assetCode": code,
            "assetId": asset_id(code),
            "assetName": clean(row.get("자산명")),
            "fundCode": clean(row.get("펀드코드")),
            "fundName": clean(row.get("펀드명")),
            "loanType": clean(row.get("대출유형")),
            "tranche": clean(row.get("Tranche")),
            "lenderName": lender,
            "drawnAmountKrw": amount(row.get("인출금액(원)")),
            "drawnAt": date_text(row.get("인출시점")),
            "maturityAt": date_text(row.get("만기시점")),
            "interestType": clean(row.get("이자유형")),
            "baseRatePct": amount(row.get("기준금리(%)")),
            "spreadRatePct": amount(row.get("가산금리(%)")),
            "loanRatePct": amount(row.get("대출금리(%)")),
            "feeRatePct": amount(row.get("수수료율(%)")),
            "allInPct": amount(row.get("All-In(%)")),
        })
    return funds, beneficiaries, lenders


def convert_staff_photos(people: list[dict[str, Any]]) -> dict[str, str]:
    source_dir = ROOT / "직원 사진"
    DOCS_STAFF.mkdir(parents=True, exist_ok=True)
    photo_map: dict[str, str] = {}
    if source_dir.exists():
        for image_path in sorted(source_dir.iterdir(), key=lambda item: item.name):
            if image_path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
                continue
            name = image_path.stem.strip()
            output = DOCS_STAFF / f"{name}.webp"
            with Image.open(image_path) as image:
                image = ImageOps.exif_transpose(image)
                image.thumbnail((320, 320))
                canvas = Image.new("RGB", image.size, "#1f1f1d")
                if image.mode == "RGBA":
                    canvas.paste(image, mask=image.getchannel("A"))
                else:
                    canvas.paste(image.convert("RGB"))
                canvas.save(output, "WEBP", quality=84, method=6)
            photo_map[name] = f"staff/{name}.webp"

    default_avatar = ROOT / "docs" / "default_avatar.svg"
    if not default_avatar.exists():
        default_avatar.write_text(
            '<svg xmlns="http://www.w3.org/2000/svg" width="128" height="128" viewBox="0 0 128 128">'
            '<rect width="128" height="128" rx="16" fill="#20201f"/>'
            '<circle cx="64" cy="48" r="24" fill="#74a86f"/>'
            '<path d="M24 112c6-26 22-40 40-40s34 14 40 40" fill="#74a86f"/>'
            '</svg>\n',
            encoding="utf-8",
        )
    for person in people:
        person["photoUrl"] = photo_map.get(person["staffName"], "default_avatar.svg")
    return photo_map


def load_current_asset_options() -> list[dict[str, Any]]:
    options = read_json(DOCS_DATA / "asset-options.json", [])
    return options if isinstance(options, list) else []


def sort_assets(options: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(options, key=lambda row: clean(row.get("assetName") or row.get("assetId")))


def new_asset_option(asset: dict[str, Any]) -> dict[str, Any]:
    return {
        "assetId": asset["assetId"],
        "assetName": asset["assetName"],
        "uniqueTenantCount": 0,
        "fetchedAt": GENERATED_AT[:10],
        "averageENoc": None,
        "vacancyRate": None,
        "monthlyCostTotal": 0,
        "dataStatus": "임대차계약 없음",
    }


def build_new_assets(
    managers: list[dict[str, Any]],
    funds: dict[str, dict[str, Any]],
    beneficiaries: list[dict[str, Any]],
    lenders: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    managers_by_code = {row["assetCode"]: row for row in managers}
    result: dict[str, dict[str, Any]] = {}
    for code in NEW_ASSET_CODES:
        fund = funds.get(code, {})
        manager = managers_by_code.get(code, {})
        override = ASSET_OVERRIDES[code]
        name = clean(manager.get("assetName") or fund.get("assetName") or override["assetName"])
        if code == "A190013001":
            name = name.strip()
        result[code] = {
            "assetCode": code,
            "assetId": asset_id(code),
            "assetName": name,
            "fundCode": clean(fund.get("fundCode") or manager.get("fundCode")),
            "fundName": clean(fund.get("fundName") or manager.get("fundName")),
            "fund": fund,
            "manager": manager,
            "beneficiaries": [row for row in beneficiaries if row["assetCode"] == code],
            "lenders": [row for row in lenders if row["assetCode"] == code],
            **override,
        }
        result[code]["assetName"] = name
    return result


def empty_asset_payload(asset: dict[str, Any], asset_options: list[dict[str, Any]]) -> dict[str, Any]:
    manager = asset.get("manager") or {}
    fund = asset.get("fund") or {}
    overview = {
        "assetId": asset["assetId"],
        "assetCode": asset["assetCode"],
        "assetName": asset["assetName"],
        "sector": "물류",
        "fundCode": asset.get("fundCode"),
        "fundName": asset.get("fundName"),
        "address": asset.get("address"),
        "standardizedAddress": asset.get("address"),
        "parcelAddress": asset.get("parcelAddress"),
        "latitude": asset.get("latitude"),
        "longitude": asset.get("longitude"),
        "coordinateStatus": asset.get("coordinateStatus"),
        "coordinateNote": asset.get("coordinateNote"),
        "manager": {
            "managerName": manager.get("managerName"),
            "teamName": manager.get("organization"),
            "organization": manager.get("organization"),
            "email": manager.get("email"),
            "assetCode": asset["assetCode"],
            "assetName": asset["assetName"],
            "fundCode": asset.get("fundCode"),
            "fundName": asset.get("fundName"),
        },
        "tenantCount": 0,
        "uniqueTenantCount": 0,
        "leaseSpaceCount": 0,
        "leasedAreaSqm": 0,
        "monthlyRentTotal": 0,
        "monthlyMfTotal": 0,
        "monthlyCostTotal": 0,
        "dataStatus": "임대차계약 없음",
        "buildingHubStatus": asset.get("buildingRegisterStatus"),
        "buildingRegisterStatus": asset.get("buildingRegisterStatus"),
        "buildingRegisterNote": asset.get("buildingRegisterNote"),
        "investmentOverview": {
            "fundType": fund.get("fundType"),
            "investmentStrategy": fund.get("investmentStrategy"),
            "legalForm": fund.get("legalForm"),
        },
        "fundOverview": fund,
        "beneficiaryCount": len(asset.get("beneficiaries") or []),
        "lenderCount": len(asset.get("lenders") or []),
    }
    return {
        "kpis": [
            {"key": "occupancy_rate", "label": "임대율", "value": None, "status": "임대차계약 없음", "valueType": "percent"},
            {"key": "leased_area_total", "label": "총 임대면적", "value": 0, "status": "임대차계약 없음", "valueType": "area"},
            {"key": "vacancy_area_total", "label": "공실면적", "value": 0, "status": "자료 없음", "valueType": "area"},
            {"key": "unique_tenant_count", "label": "현재 임차인 수", "value": 0, "status": "임대차계약 없음", "valueType": "number"},
            {"key": "average_rent_per_py", "label": "평당 임대료 평균", "value": "자료 없음", "status": "임대차계약 없음"},
            {"key": "average_mf_per_py", "label": "평당 관리비 평균", "value": "자료 없음", "status": "임대차계약 없음"},
            {"key": "average_e_noc", "label": "E.NOC", "value": "자료 없음", "status": "임대차계약 없음"},
        ],
        "meta": {
            "page": "asset",
            "rowCount": 0,
            "selection": {"assetId": asset["assetId"], "assetName": asset["assetName"]},
            "generatedAt": GENERATED_AT,
            "basisDisplay": {
                "asOf": GENERATED_AT[:10],
                "page": "asset",
                "source": "260513/260520 Excel supplemental",
                "generatedAt": GENERATED_AT,
                "refreshedAt": GENERATED_AT,
            },
        },
        "rows": [],
        "basis": {
            "page": "asset",
            "rowScope": "supplemental_new_asset_without_lease",
            "nullPolicy": "신규 자산의 임대차계약 데이터가 없으면 금액/면적 평균은 자료 없음으로 표시합니다.",
        },
        "filters": {"assets": asset_options, "selectedAssetId": asset["assetId"]},
        "overview": overview,
        "analytics": {"coreTenants": [], "contractExpiry": [], "monthlyCostByTenant": []},
        "topTenants": [],
        "stackingPlan": [],
        "areaBreakdown": {
            "leasedAreaSqm": 0,
            "grossFloorAreaSqm": None,
            "vacancyAreaSqm": None,
            "dataStatus": "임대차계약 없음",
        },
        "fundOverview": fund,
        "beneficiaries": asset.get("beneficiaries") or [],
        "lenders": asset.get("lenders") or [],
        "generatedAt": GENERATED_AT,
        "basisDisplay": {
            "asOf": GENERATED_AT[:10],
            "page": "asset",
            "source": "260513/260520 Excel supplemental",
            "generatedAt": GENERATED_AT,
            "refreshedAt": GENERATED_AT,
        },
        "schemaVersion": "docs_static_supplemental_v1",
        "payloadSource": "supabase_snapshot",
        "dataSourceMode": "supabase_snapshot",
        "sourceSystem": "excel_supplemental",
    }


def ensure_row_by_asset(rows: list[dict[str, Any]], asset: dict[str, Any], factory) -> list[dict[str, Any]]:
    rows = [row for row in rows if row.get("assetId") != asset["assetId"] and row.get("assetCode") != asset["assetCode"]]
    rows.append(factory(asset))
    return rows


def sanitize_public_admin_payload(admin_payload: dict[str, Any]) -> dict[str, Any]:
    public_payload = json.loads(json.dumps(admin_payload, ensure_ascii=False))
    for row in public_payload.get("staffProfiles", []):
        row["email"] = ""
    for row in public_payload.get("userPermissions", []):
        row["email"] = ""
    for row in public_payload.get("assetManagers", []):
        row["email"] = ""
    public_payload["loginHistory"] = []
    public_payload["publicFallbackNotice"] = "Static GitHub fallback omits staff email and persisted login history. Supabase Edge Function provides live admin history when an auth session is present."
    return public_payload


def apply_docs_data(new_assets: dict[str, dict[str, Any]], admin_payload: dict[str, Any]) -> None:
    asset_options = load_current_asset_options()
    existing_by_id = {row.get("assetId"): row for row in asset_options}
    for asset in new_assets.values():
        existing_by_id[asset["assetId"]] = {**existing_by_id.get(asset["assetId"], {}), **new_asset_option(asset)}
    asset_options = sort_assets([row for row in existing_by_id.values() if row.get("assetId")])
    write_json(DOCS_DATA / "asset-options.json", asset_options)

    for asset in new_assets.values():
        write_json(DOCS_DATA / "asset" / f"{asset['assetId']}.json", empty_asset_payload(asset, asset_options))

    bootstrap = read_json(DOCS_DATA / "bootstrap.json", {})
    bootstrap["assetOptions"] = asset_options
    bootstrap["supplementalAssets"] = list(new_assets.values())
    bootstrap["adminSummary"] = {
        "permissionRows": len(admin_payload.get("userPermissions", [])),
        "staffProfiles": len(admin_payload.get("staffProfiles", [])),
        "newAssetCodes": NEW_ASSET_CODES,
        "generatedAt": GENERATED_AT,
    }
    bootstrap["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "bootstrap.json", bootstrap)

    home = read_json(DOCS_DATA / "home.json", {})
    home_points = home.get("mapPoints") if isinstance(home.get("mapPoints"), list) else []
    home_vacancy = home.get("vacancySummary") if isinstance(home.get("vacancySummary"), list) else []
    for asset in new_assets.values():
        def map_factory(item):
            return {
                "assetId": item["assetId"],
                "assetCode": item["assetCode"],
                "assetName": item["assetName"],
                "address": item.get("address"),
                "latitude": item.get("latitude"),
                "longitude": item.get("longitude"),
                "coordinateStatus": item.get("coordinateStatus"),
                "coordinateNote": item.get("coordinateNote"),
                "issueCount": 0,
                "dataStatus": "임대차계약 없음",
                "buildingRegisterStatus": item.get("buildingRegisterStatus"),
            }
        home_points = ensure_row_by_asset(home_points, asset, map_factory)
        home_vacancy = ensure_row_by_asset(home_vacancy, asset, lambda item: {
            "assetId": item["assetId"],
            "assetName": item["assetName"],
            "vacancyRate": None,
            "vacancyAreaSqm": None,
            "grossFloorAreaSqm": None,
            "dataStatus": "임대차계약 없음",
        })
    home["mapPoints"] = sorted(home_points, key=lambda row: clean(row.get("assetName")))
    home["vacancySummary"] = sorted(home_vacancy, key=lambda row: clean(row.get("assetName")))
    home["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "home.json", home)

    sector = read_json(DOCS_DATA / "sector.json", {})
    sector_points = sector.get("mapPoints") if isinstance(sector.get("mapPoints"), list) else []
    for asset in new_assets.values():
        sector_points = ensure_row_by_asset(sector_points, asset, lambda item: {
            "assetId": item["assetId"],
            "assetCode": item["assetCode"],
            "assetName": item["assetName"],
            "address": item.get("address"),
            "latitude": item.get("latitude"),
            "longitude": item.get("longitude"),
            "coordinateStatus": item.get("coordinateStatus"),
            "coordinateNote": item.get("coordinateNote"),
            "issueCount": 0,
            "dataStatus": "임대차계약 없음",
        })
    sector["mapPoints"] = sorted(sector_points, key=lambda row: clean(row.get("assetName")))
    sector["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "sector.json", sector)

    weekly = read_json(DOCS_DATA / "weekly.json", {})
    asset_rows = weekly.get("assetRows") if isinstance(weekly.get("assetRows"), list) else []
    for index, asset in enumerate(new_assets.values(), start=1):
        asset_rows = ensure_row_by_asset(asset_rows, asset, lambda item, idx=index: {
            "id": item["assetId"],
            "no": 900 + idx,
            "category": "신규 자산",
            "assetId": item["assetId"],
            "assetCode": item["assetCode"],
            "assetName": item["assetName"],
            "fundName": item.get("fundName"),
            "fundMaturity": item.get("fund", {}).get("maturityDate"),
            "loanMaturity": max([row.get("maturityAt") or "" for row in item.get("lenders", [])] or [""]),
            "mainTenant": "임대차계약 없음",
            "mainIssue": item.get("buildingRegisterNote"),
            "completion": "개발 중" if item["assetCode"] == "A190013001" else "",
            "managerName": item.get("manager", {}).get("managerName"),
        })
    weekly["assetRows"] = sorted(asset_rows, key=lambda row: clean(row.get("assetName")))
    weekly["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "weekly.json", weekly)

    source_stub_rows = []
    for asset in new_assets.values():
        source_stub_rows.append({
            "sector": "물류",
            "assetId": asset["assetId"],
            "assetCode": asset["assetCode"],
            "fundCode": asset.get("fundCode"),
            "fundName": asset.get("fundName"),
            "tenantId": "",
            "tenantMasterName": "임대차계약 없음",
            "assetName": asset["assetName"],
            "leasedAreaSqm": 0,
            "leasedAreaPy": 0,
            "monthlyRentTotal": 0,
            "monthlyMfTotal": 0,
            "monthlyCostTotal": 0,
            "dataStatus": "임대차계약 없음",
        })

    tools = read_json(DOCS_DATA / "tools.json", {})
    for key in ["assets", "benchmarkRows"]:
        rows = tools.get(key) if isinstance(tools.get(key), list) else []
        for stub in source_stub_rows:
            rows = ensure_row_by_asset(rows, stub, lambda item: item.copy())
        tools[key] = sorted(rows, key=lambda row: clean(row.get("assetName")))
    for key in ["sourceRows"]:
        rows = tools.get(key) if isinstance(tools.get(key), list) else []
        for stub in source_stub_rows:
            rows = ensure_row_by_asset(rows, stub, lambda item: item.copy())
        tools[key] = sorted(rows, key=lambda row: clean(row.get("assetName")))
    tools["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "tools.json", tools)

    playground = read_json(DOCS_DATA / "playground.json", {})
    rows = playground.get("sourceRows") if isinstance(playground.get("sourceRows"), list) else []
    for stub in source_stub_rows:
        rows = ensure_row_by_asset(rows, stub, lambda item: item.copy())
    playground["sourceRows"] = sorted(rows, key=lambda row: clean(row.get("assetName")))
    playground["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "playground.json", playground)

    public_admin_payload = sanitize_public_admin_payload(admin_payload)
    public_admin_payload["generatedAt"] = GENERATED_AT
    public_admin_payload["schemaVersion"] = "docs_static_admin_supplemental_v1"
    public_admin_payload["payloadSource"] = "supabase_snapshot"
    public_admin_payload["dataSourceMode"] = "supabase_snapshot"
    public_admin_payload["sourceSystem"] = "excel_supplemental"
    write_json(DOCS_DATA / "admin.json", public_admin_payload)

    initial = read_json(DOCS_DATA / "initial.json", {})
    initial["tabPayloads"] = {
        "weekly": read_json(DOCS_DATA / "weekly.json", {}),
        "home": read_json(DOCS_DATA / "home.json", {}),
        "sector": read_json(DOCS_DATA / "sector.json", {}),
        "tools": read_json(DOCS_DATA / "tools.json", {}),
        "playground": read_json(DOCS_DATA / "playground.json", {}),
    }
    initial["generatedAt"] = GENERATED_AT
    write_json(DOCS_DATA / "initial.json", initial)


def build_admin_payload(
    people: list[dict[str, Any]],
    managers: list[dict[str, Any]],
    new_assets: dict[str, dict[str, Any]],
    beneficiaries: list[dict[str, Any]],
    lenders: list[dict[str, Any]],
) -> dict[str, Any]:
    asset_name_by_code = {row["assetCode"]: row["assetName"].strip() for row in managers}
    permissions = []
    for person in sorted(people, key=lambda row: row["staffName"]):
        asset_names = [asset_name_by_code.get(code, code).strip() for code in person["assetCodes"]]
        permissions.append({
            **person,
            "assetNames": asset_names,
            "assetCount": len(person["assetCodes"]),
            "assignedRead": "Y" if person["assignedCanRead"] else "N",
            "assignedCreate": "Y" if person["assignedCanCreate"] else "N",
            "assignedUpdate": "Y" if person["assignedCanUpdate"] else "N",
            "assignedDelete": "Y" if person["assignedCanDelete"] else "N",
            "otherRead": "Y" if person["otherCanRead"] else "N",
            "otherCreate": "Y" if person["otherCanCreate"] else "N",
            "otherUpdate": "Y" if person["otherCanUpdate"] else "N",
            "otherDelete": "Y" if person["otherCanDelete"] else "N",
        })

    staff_profiles = [{
        "staffId": safe_id(person["email"] or person["staffName"]),
        "staffName": person["staffName"],
        "email": person["email"],
        "organization": person["organization"],
        "photoUrl": person.get("photoUrl", "default_avatar.svg"),
    } for person in sorted(people, key=lambda row: row["staffName"])]

    new_asset_codes = set(new_assets)
    return {
        "kpis": [
            {"key": "staff_count", "label": "권한 사용자 수", "value": len(people), "status": "ok", "valueType": "number"},
            {"key": "permission_asset_count", "label": "권한 자산 수", "value": len(set(code for p in people for code in p["assetCodes"])), "status": "ok", "valueType": "number"},
            {"key": "new_asset_count", "label": "신규 자산", "value": len(new_assets), "status": "ok", "valueType": "number"},
            {"key": "apps_script_calls", "label": "Apps Script 호출", "value": 0, "status": "blocked by QA", "valueType": "number"},
        ],
        "runtime": {
            "dataSourceMode": "supabase_snapshot",
            "payloadSource": "supabase_snapshot",
            "screenRuntime": "GitHub Pages static frontend",
            "externalRuntimeBridge": "Supabase Edge Function only",
            "appsScriptRuntime": "legacy reference only",
            "frontendSecretKey": "none",
        },
        "staffProfiles": staff_profiles,
        "userPermissions": permissions,
        "assetManagers": sorted(managers, key=lambda row: clean(row.get("assetName"))),
        "newAssets": list(new_assets.values()),
        "fundBeneficiaries": [row for row in beneficiaries if row["assetCode"] in new_asset_codes],
        "fundLenders": [row for row in lenders if row["assetCode"] in new_asset_codes],
        "loginHistory": [
            {
                "eventAt": GENERATED_AT,
                "staffName": "System",
                "email": "",
                "eventType": "snapshot_generated",
                "status": "ready",
                "source": "docs/data/admin.json",
            }
        ],
    }


def source_ref(row: dict[str, Any], sheet: str) -> str:
    return f"{row.get('sourceWorkbook', 'excel')}!{sheet}!{row.get('sourceRow', '')}"


def build_supplemental_dataset(
    people: list[dict[str, Any]],
    managers: list[dict[str, Any]],
    new_assets: dict[str, dict[str, Any]],
    admin_payload: dict[str, Any],
) -> dict[str, Any]:
    tables: dict[str, list[dict[str, Any]]] = {
        "ll_funds": [],
        "ll_assets": [],
        "ll_asset_managers": [],
        "ll_user_permissions": [],
        "ll_staff_profiles": [],
        "ll_fund_beneficiaries": [],
        "ll_fund_lenders": [],
        "ll_login_history": [],
    }

    for asset in new_assets.values():
        fund = asset.get("fund") or {}
        f_id = fund_id(asset.get("fundCode") or asset["assetCode"])
        fund_payload = {**fund, "source": "260520_물류센터 펀드 정보.xlsx"}
        tables["ll_funds"].append({
            "fund_id": f_id,
            "fund_code": asset.get("fundCode"),
            "fund_name": asset.get("fundName"),
            "raw_fund_name": asset.get("fundName"),
            "short_name": fund.get("shortName"),
            "sector": fund.get("sector") or "물류",
            "setup_date": fund.get("setupDate") or None,
            "maturity_date": fund.get("maturityDate") or None,
            "status": "active",
            "source_system": "google_sheets",
            "source_table": "260520_펀드 정보",
            "source_pk": asset["assetCode"],
            "source_ref": source_ref(fund, "펀드 정보"),
            "source_row_hash": stable_hash(fund_payload),
            "source_payload": fund_payload,
            "review_status": "ok",
        })
        asset_payload = {
            "address": asset.get("address"),
            "parcelAddress": asset.get("parcelAddress"),
            "coordinateStatus": asset.get("coordinateStatus"),
            "coordinateNote": asset.get("coordinateNote"),
            "buildingRegisterStatus": asset.get("buildingRegisterStatus"),
            "buildingRegisterNote": asset.get("buildingRegisterNote"),
            "buildingRegisterQuery": asset.get("buildingRegisterQuery"),
            "fund": fund,
        }
        tables["ll_assets"].append({
            "asset_id": asset["assetId"],
            "asset_code": asset["assetCode"],
            "asset_name": asset["assetName"],
            "raw_asset_name": asset["assetName"],
            "fund_id": f_id,
            "sector": "물류",
            "address": asset.get("address"),
            "latitude": asset.get("latitude"),
            "longitude": asset.get("longitude"),
            "source_system": "google_sheets",
            "source_table": "260520_펀드 정보",
            "source_pk": asset["assetCode"],
            "source_ref": source_ref(fund, "펀드 정보"),
            "source_row_hash": stable_hash(asset_payload),
            "source_payload": asset_payload,
            "review_status": "review_required" if "not_found" in asset.get("buildingRegisterStatus", "") else "ok",
            "review_note": asset.get("buildingRegisterNote"),
        })

    for manager in managers:
        if manager["assetCode"] not in NEW_ASSET_CODES:
            continue
        payload = {**manager, "source": "260513_담당자별 권한 부여_수식 제거.xlsx"}
        tables["ll_asset_managers"].append({
            "asset_manager_id": "asset_manager_" + safe_id(manager["assetCode"]),
            "asset_id": manager["assetId"],
            "asset_code": manager["assetCode"],
            "asset_name": manager["assetName"].strip(),
            "fund_id": fund_id(manager["fundCode"]),
            "fund_code": manager["fundCode"],
            "fund_name": manager["fundName"],
            "manager_name": manager["managerName"],
            "organization": manager["organization"],
            "email": manager["email"],
            "source_system": "google_sheets",
            "source_table": "260513_자산별 담당자",
            "source_pk": manager["assetCode"],
            "source_ref": source_ref(manager, "담당자별 권한 부여"),
            "source_row_hash": stable_hash(payload),
            "source_payload": payload,
        })

    for permission in admin_payload["userPermissions"]:
        tables["ll_staff_profiles"].append({
            "staff_id": safe_id(permission["email"] or permission["staffName"]),
            "staff_name": permission["staffName"],
            "email": permission["email"],
            "organization": permission["organization"],
            "photo_url": permission.get("photoUrl"),
            "source_system": "google_sheets",
            "source_table": "260513_인원별 권한",
            "source_pk": permission["email"] or permission["staffName"],
            "source_ref": source_ref(permission, "담당자별 권한 부여"),
            "source_row_hash": stable_hash(permission),
            "source_payload": permission,
        })
        for code in permission["assetCodes"]:
            scope = asset_id(code)
            payload = {"assetCode": code, **permission}
            tables["ll_user_permissions"].append({
                "permission_id": permission_id(permission["email"] or permission["staffName"], "asset", scope),
                "principal_type": "user_email",
                "principal_id": permission["email"] or permission["staffName"],
                "scope_type": "asset",
                "scope_id": scope,
                "can_read": permission["assignedCanRead"],
                "can_write": permission["assignedCanCreate"] or permission["assignedCanUpdate"],
                "can_delete": permission["assignedCanDelete"],
                "source_system": "google_sheets",
                "created_by": "260513_담당자별 권한 부여_수식 제거.xlsx",
                "source_payload": payload,
            })
        if permission["otherCanRead"] or permission["otherCanCreate"] or permission["otherCanUpdate"] or permission["otherCanDelete"]:
            payload = {"scope": "other_assets", **permission}
            tables["ll_user_permissions"].append({
                "permission_id": permission_id(permission["email"] or permission["staffName"], "other_assets", "all"),
                "principal_type": "user_email",
                "principal_id": permission["email"] or permission["staffName"],
                "scope_type": "other_assets",
                "scope_id": "all",
                "can_read": permission["otherCanRead"],
                "can_write": permission["otherCanCreate"] or permission["otherCanUpdate"],
                "can_delete": permission["otherCanDelete"],
                "source_system": "google_sheets",
                "created_by": "260513_담당자별 권한 부여_수식 제거.xlsx",
                "source_payload": payload,
            })

    for row in admin_payload["fundBeneficiaries"]:
        payload = {**row, "source": "260520_물류센터 펀드 정보.xlsx"}
        tables["ll_fund_beneficiaries"].append({
            "beneficiary_id": "beneficiary_" + safe_id(f"{row['assetCode']}_{row['beneficiaryName']}_{row['sourceRow']}"),
            "asset_id": row["assetId"],
            "asset_code": row["assetCode"],
            "fund_id": fund_id(row["fundCode"]),
            "fund_code": row["fundCode"],
            "fund_name": row["fundName"],
            "tranche": row["tranche"],
            "beneficiary_name": row["beneficiaryName"],
            "investment_amount_krw": row["investmentAmountKrw"],
            "source_system": "google_sheets",
            "source_table": "260520_수익자 정보",
            "source_pk": f"{row['assetCode']}|{row['sourceRow']}",
            "source_ref": source_ref(row, "수익자 정보"),
            "source_row_hash": stable_hash(payload),
            "source_payload": payload,
        })

    for row in admin_payload["fundLenders"]:
        payload = {**row, "source": "260520_물류센터 펀드 정보.xlsx"}
        tables["ll_fund_lenders"].append({
            "lender_id": "lender_" + safe_id(f"{row['assetCode']}_{row['lenderName']}_{row['sourceRow']}"),
            "asset_id": row["assetId"],
            "asset_code": row["assetCode"],
            "fund_id": fund_id(row["fundCode"]),
            "fund_code": row["fundCode"],
            "fund_name": row["fundName"],
            "loan_type": row["loanType"],
            "tranche": row["tranche"],
            "lender_name": row["lenderName"],
            "drawn_amount_krw": row["drawnAmountKrw"],
            "drawn_at": row["drawnAt"] or None,
            "maturity_at": row["maturityAt"] or None,
            "interest_type": row["interestType"],
            "base_rate_pct": row["baseRatePct"],
            "spread_rate_pct": row["spreadRatePct"],
            "loan_rate_pct": row["loanRatePct"],
            "fee_rate_pct": row["feeRatePct"],
            "all_in_pct": row["allInPct"],
            "source_system": "google_sheets",
            "source_table": "260520_대주 정보",
            "source_pk": f"{row['assetCode']}|{row['sourceRow']}",
            "source_ref": source_ref(row, "대주 정보"),
            "source_row_hash": stable_hash(payload),
            "source_payload": payload,
        })

    for row in admin_payload["loginHistory"]:
        tables["ll_login_history"].append({
            "login_event_id": "login_event_" + stable_hash(row)[:16],
            "event_at": row["eventAt"],
            "staff_name": row["staffName"],
            "email": row["email"],
            "event_type": row["eventType"],
            "status": row["status"],
            "source_payload": row,
        })

    return {"generatedAt": GENERATED_AT, "tables": tables}


def main() -> None:
    people, managers = parse_permission_workbook()
    funds, beneficiaries, lenders = parse_fund_workbook()
    convert_staff_photos(people)
    new_assets = build_new_assets(managers, funds, beneficiaries, lenders)
    admin_payload = build_admin_payload(people, managers, new_assets, beneficiaries, lenders)
    apply_docs_data(new_assets, admin_payload)

    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    supplemental = build_supplemental_dataset(people, managers, new_assets, admin_payload)
    write_json(ARTIFACT_DIR / "logistics-supplemental-dataset.json", supplemental)
    write_json(ARTIFACT_DIR / "logistics-supplemental-summary.json", {
        "generatedAt": GENERATED_AT,
        "newAssets": [
            {
                "assetCode": asset["assetCode"],
                "assetId": asset["assetId"],
                "assetName": asset["assetName"],
                "manager": asset.get("manager", {}).get("managerName"),
                "beneficiaries": len(asset.get("beneficiaries") or []),
                "lenders": len(asset.get("lenders") or []),
                "buildingRegisterStatus": asset.get("buildingRegisterStatus"),
            }
            for asset in new_assets.values()
        ],
        "staffCount": len(people),
        "permissionRows": len(supplemental["tables"]["ll_user_permissions"]),
        "photoCount": len(list(DOCS_STAFF.glob("*.webp"))),
        "tables": {name: len(rows) for name, rows in supplemental["tables"].items()},
    })
    print(json.dumps({
        "ok": True,
        "generatedAt": GENERATED_AT,
        "newAssets": list(new_assets),
        "staffCount": len(people),
        "permissionRows": len(supplemental["tables"]["ll_user_permissions"]),
        "photoCount": len(list(DOCS_STAFF.glob("*.webp"))),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
